from hey.mcp_tools.camel.toolkits.base import BaseToolkit
from hey.mcp_tools.camel.toolkits.function_tool import FunctionTool
from retry import retry
from typing import List, Dict, Any, Optional, Tuple
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tabulate import tabulate
from xls2xlsx import XLS2XLSX
import os
import pandas as pd


class ExcelToolkit(BaseToolkit):
    r"""A class representing a toolkit for extract detailed cell information from an Excel file.

    This class provides method for processing docx, pdf, pptx, etc. It cannot process excel files.
    """

    def _convert_to_markdown(self, df: pd.DataFrame) -> str:
        """
        Convert DataFrame to Markdown format table.
        
        Args:
            df (pd.DataFrame): DataFrame containing the Excel data.
        
        Returns:
            str: Markdown formatted table.
        """
        md_table = tabulate(df, headers='keys', tablefmt='pipe')
        return str(md_table)
    

    def extract_excel_content(self, document_path: str) -> str:
        r"""Extract detailed cell information from an Excel file, including multiple sheets.
        
        Args:
            document_path (str): The path of the Excel file.
        
        Returns:
            str: Extracted excel information, including details of each sheet.
        """
        logger.debug(f"Calling extract_excel_content with document_path: {document_path}")

        if not (document_path.endswith("xls") or document_path.endswith("xlsx") or document_path.endswith("csv")):
            logger.error("Only xls, xlsx, csv files are supported.")
            return f"Failed to process file {document_path}: It is not excel format. Please try other ways."

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"


        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        # Load the Excel workbook
        wb = load_workbook(document_path, data_only=True)
        sheet_info_list = [] 

        # Iterate through all sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            cell_info_list = []

            for row in ws.iter_rows():
                for cell in row:
                    row_num = cell.row
                    col_letter = cell.column_letter

                    cell_value = cell.value

                    font_color = None
                    if cell.font and cell.font.color and "rgb=None" not in str(cell.font.color):  # Handle font color
                        font_color = cell.font.color.rgb

                    fill_color = None
                    if cell.fill and cell.fill.fgColor and "rgb=None" not in str(cell.fill.fgColor):  # Handle fill color
                        fill_color = cell.fill.fgColor.rgb

                    cell_info_list.append({
                        "index": f"{row_num}{col_letter}",
                        "value": cell_value,
                        "font_color": font_color,
                        "fill_color": fill_color,
                    })

            # Convert the sheet to a DataFrame and then to markdown
            sheet_df = pd.read_excel(document_path, sheet_name=sheet, engine='openpyxl')
            markdown_content = self._convert_to_markdown(sheet_df)

            # Collect all information for the sheet
            sheet_info = {
                "sheet_name": sheet,
                "cell_info_list": cell_info_list,
                "markdown_content": markdown_content,
            }
            sheet_info_list.append(sheet_info)

        result_str = ""
        for sheet_info in sheet_info_list:
            result_str += f"""
            Sheet Name: {sheet_info['sheet_name']}
            Cell information list:
            {sheet_info['cell_info_list']}
            
            Markdown View of the content:
            {sheet_info['markdown_content']}
            
            {'-'*40}
            """

        return result_str

    def get_tools(self) -> List[FunctionTool]:
        r"""Returns a list of FunctionTool objects representing the functions in the toolkit.

        Returns:
            List[FunctionTool]: A list of FunctionTool objects representing the functions in the toolkit.
        """
        return [
            FunctionTool(self.extract_excel_content),
        ]
